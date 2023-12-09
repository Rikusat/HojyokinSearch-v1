import streamlit as st
from openpyxl import load_workbook
import pandas as pd
import io
import os
from docx import Document
import io
import streamlit as st

def generate_document():
    # 新しいWordドキュメントの作成
    doc = Document()
    doc.add_paragraph("Hello, World!")  # 例としてテキストを追加

    # バイナリデータに変換
    doc_buffer = io.BytesIO()
    doc.save(doc_buffer)
    doc_bytes = doc_buffer.getvalue()

    return doc_bytes

# 編集されたWordドキュメントの生成
edited_doc = generate_document()

# ダウンロードボタンの表示
st.download_button(label='Download Edits', data=edited_doc, file_name='EDITED.docx', mime='application/octet-stream', key=321)


def replace_text_in_word(input_word_file, output_word_file, replacements):
    with open(input_word_file, 'rb') as file:
        doc_bytes = io.BytesIO(file.read())

    doc_text = doc_bytes.getvalue().decode("utf-8")

    for old_text, new_text in replacements.items():
        doc_text = doc_text.replace(old_text, new_text)

    with open(output_word_file, 'wb') as file:
        file.write(doc_text.encode("utf-8"))

def display_excel_table(excel_file):
    df = pd.read_excel(excel_file)
    st.subheader("Excelファイルの内容:")
    st.write(df)

def main():
    st.title('Word書類の文字列置換アプリ')

    # Excelファイルのアップロード
    st.sidebar.header('Excelファイルをアップロード')
    excel_file = st.sidebar.file_uploader("Excelファイルを選択してください", type=['xlsx'])

    # Wordファイルのアップロード
    st.sidebar.header('Wordファイルをアップロード')
    word_files = st.sidebar.file_uploader("Wordファイルを選択してください", type=['docx'], accept_multiple_files=True)

    if excel_file and word_files:
        # Excelファイルの一時保存
        excel_bytes = excel_file.read()
        excel_path = os.path.join("./", "temp_excel.xlsx")
        with open(excel_path, "wb") as temp_excel:
            temp_excel.write(excel_bytes)

        # Excelファイルから置換情報を取得
        wb = load_workbook(excel_path)
        ws = wb.active

        replacements = {}
        max_col = ws.max_column
        max_row = ws.max_row

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                old_text = ws.cell(row=row, column=col).value
                new_text = ws.cell(row=row, column=col + 1).value

                if old_text is not None and new_text is not None:
                    replacements[old_text] = new_text

        # Wordファイルごとに処理
        for word_file in word_files:
            # Wordファイルの一時保存
            word_bytes = word_file.read()
            word_path = os.path.join("./", f"temp_word_{word_file.name}")
            with open(word_path, "wb") as temp_word:
                temp_word.write(word_bytes)

            # Wordファイルの置換
            replace_text_in_word(word_path, f"output_{word_file.name}", replacements)

            # ダウンロードリンクの作成
            with open(f"output_{word_file.name}", "rb") as file:
                file_contents = file.read()
                st.sidebar.markdown(get_binary_file_downloader_html(file_contents, file_name=f"output_{word_file.name}"), unsafe_allow_html=True)

        st.success("置換が完了しました。")

# 他の関数もそのまま残す...

if __name__ == '__main__':
    main()

